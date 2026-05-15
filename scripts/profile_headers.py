#!/usr/bin/env python3
"""List column names for each raw PERM file/sheet."""

from __future__ import annotations

import argparse
import csv
import io
import json
import re
import zipfile
from pathlib import Path

from openpyxl import load_workbook


def read_csv_headers(content: bytes) -> list[str]:
    text = content.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        if row and any(cell.strip() for cell in row):
            return [cell.strip() for cell in row if cell and cell.strip()]
    return []


def read_xlsx_sheets(path_or_bytes: str | bytes, from_bytes: bool = False) -> dict[str, list[str]]:
    if from_bytes:
        wb = load_workbook(io.BytesIO(path_or_bytes), read_only=True, data_only=True)
    else:
        wb = load_workbook(path_or_bytes, read_only=True, data_only=True)

    out: dict[str, list[str]] = {}
    for ws in wb.worksheets:
        headers: list[str] = []
        for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
            values = ["" if v is None else str(v).strip() for v in row]
            non_empty = [v for v in values if v]
            if len(non_empty) < 3:
                continue

            mostly_text = sum(1 for v in non_empty if not re.fullmatch(r"[0-9.,-]+", v)) >= max(2, len(non_empty) // 2)
            if mostly_text:
                headers = [v for v in values if v]
                break

        out[ws.title] = headers

    wb.close()
    return out


def profile_file(path: Path) -> dict:
    ext = path.suffix.lower()
    rec = {
        "file": path.name,
        "type": ext,
        "entries": [],
    }

    def add_entry(container: str, headers: list[str]) -> None:
        rec["entries"].append(
            {
                "container": container,
                "column_names": headers,
            }
        )

    if ext == ".xlsx":
        sheets = read_xlsx_sheets(str(path))
        for sheet, headers in sheets.items():
            add_entry(f"sheet:{sheet}", headers)
    elif ext == ".csv":
        add_entry("csv", read_csv_headers(path.read_bytes()))
    elif ext == ".zip":
        with zipfile.ZipFile(path, "r") as zf:
            for name in zf.namelist():
                lower = name.lower()
                if lower.endswith(".csv"):
                    add_entry(f"zip:{name}", read_csv_headers(zf.read(name)))
                elif lower.endswith(".xlsx"):
                    sheets = read_xlsx_sheets(zf.read(name), from_bytes=True)
                    for sheet, headers in sheets.items():
                        add_entry(f"zip:{name}#sheet:{sheet}", headers)

    return rec


def main() -> None:
    parser = argparse.ArgumentParser(description="List column names across raw PERM files")
    parser.add_argument("--raw-dir", default="data/raw", help="Directory of raw files")
    parser.add_argument("--out-json", default="data/processed/header_profile.json", help="Output JSON path")
    parser.add_argument("--out-md", default="data/processed/header_profile.md", help="Output markdown summary path")
    args = parser.parse_args()

    raw_dir = Path(args.raw_dir)
    out_json = Path(args.out_json)
    out_md = Path(args.out_md)
    out_json.parent.mkdir(parents=True, exist_ok=True)

    files = sorted([p for p in raw_dir.glob("*") if p.suffix.lower() in {".xlsx", ".csv", ".zip"}])
    profiles = [profile_file(p) for p in files]

    payload = {
        "files_profiled": len(profiles),
        "profiles": profiles,
    }
    out_json.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    lines: list[str] = []
    lines.append("# PERM Column Names")
    lines.append("")
    lines.append(f"Files profiled: **{len(profiles)}**")
    lines.append("")

    for pf in profiles:
        lines.append(f"## {pf['file']}")
        if not pf["entries"]:
            lines.append("- No readable tabular entries")
            lines.append("")
            continue

        for e in pf["entries"]:
            lines.append(f"- {e['container']}")
            cols = e["column_names"]
            if not cols:
                lines.append("  - (no headers detected)")
            else:
                for c in cols:
                    lines.append(f"  - {c}")
        lines.append("")

    out_md.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(f"Profiled files: {len(profiles)}")
    print(f"JSON: {out_json}")
    print(f"Markdown: {out_md}")


if __name__ == "__main__":
    main()
